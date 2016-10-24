from openpyxl import load_workbook


class Excel_Reader(object):

    def __init__(self, excel_file, sheet_name='SWITCHES', variable_name_row=1,
                 key_column=1, template_column=2, variable_start_column=3):
        ''' Loads the excel configuration file '''
        try:
            # TODO: Check is file is locked, if so wait and try to open N
            # more times after N seconds.
            self._wb = load_workbook(excel_file, data_only=True)
            self._filename = excel_file
        except:
            raise Exception("Could not load spreadsheet '%s'" % excel_file)

        self.set_excel_sheet(sheet_name)
        self.set_variable_name_row(variable_name_row)
        self.set_template_column(template_column)
        self.set_data_start_row(variable_name_row + 1)
        self.set_key_column(key_column)
        self.set_variable_start_column(variable_start_column)

    def save(self):
        ''' Saves spreadsheet to disk '''
        self._wb.save(self._filename)

    def set_data_start_row(self, data_start_row):
        ''' Set the row in the sheet where the data starts '''
        self._data_start_row = data_start_row

    def set_variable_name_row(self, variable_name_row):
        ''' Set the row where the variable names are '''
        self._variable_name_row = variable_name_row

    def set_key_column(self, key_column):
        ''' Set the key column in the template '''
        self._key_column = key_column
        self._keys = {}
        r = self._data_start_row
        while True:
            key = self._ws.cell(column=self._key_column, row=r)
            if key.value:
                # Check for duplicate key
                if key.value not in self._keys:
                    self._keys[key.value] = r
                else:
                    raise Exception("Duplicate key '%s' found at row '%s'." % (key.value, r))
            else:
                self._data_end_row = r
                break
            r += 1

    def set_template_column(self, template_column):
        ''' Set the template column in the template '''
        self._template_column = template_column

    def set_variable_start_column(self, variable_start_column):
        ''' Set the template column in the template '''
        c = self._variable_start_column = variable_start_column
        while True:
            cell = self._ws.cell(column=c, row=self._variable_name_row)
            if cell.value:
                c += 1
            else:
                self._variable_end_column = c
                break

    def set_excel_sheet(self, sheet_name):
        ''' Sets the current sheet '''
        self._sheet_name = sheet_name
        try:
            self._ws = self._wb.get_sheet_by_name(self._sheet_name)
        except:
            self._ws = self._wb.create_sheet(self._sheet_name)

        # TODO: re-set all the variables

    def get_row_for_key(self, key):
        ''' Returns the row for a given key, or -1 if not matched '''
        if key in self._keys:
            return self._keys[key]
        return -1

    def get_variables_for_key(self, key):
        ''' Returns the variables for a particular key '''
        variables = {}
        col = self._variable_start_column
        row = self.get_row_for_key(key)
        while row >= 0 and col <= self._variable_end_column:
            variable = self._ws.cell(column=col, row=self._variable_name_row)
            if variable.value:
                variables[variable.value] = self._ws.cell(column=col, row=row).value
            col += 1
        return variables

    def get_template_name_for_key(self, key):
        ''' Returns the template to file use '''
        row = self.get_row_for_key(key)
        if row >= 0:
            return self._ws.cell(column=self._template_column, row=row).value
        return ""

    def get_last_row(self):
        return self._data_end_row

    def set_values_for_variables(self, key, dict):
        # TODO: Need to probably lock file!

        assert(key in dict)

        # Build a dictionary of variables in the spreadsheet
        variables = {}
        col = self._variable_start_column
        while True:
            variable = self._ws.cell(column=col, row=self._variable_name_row)
            if variable.value:
                variables[variable.value] = col
            else:
                self._variable_end_column = col
                break
            col += 1

        # Find row for key
        print("Find row for key....")
        row = self.get_row_for_key(dict[key])
        if row == -1:
            if self._data_end_row == self._data_start_row:
                self._ws.cell(column=self._key_column, row=self._variable_name_row).value = key
            row = self._data_end_row
            self._ws.cell(column=self._key_column, row=row).value = dict[key]
            self._data_end_row += 1

        # Fill in values
        for k, v in dict.items():
            if not k == key:
                if k in variables:
                    self._ws.cell(column=variables[k], row=row).value = v
                else:
                    # Add a new column at the end with variable
                    self._ws.cell(column=self._variable_end_column,
                                  row=self._variable_name_row).value = k
                    self._ws.cell(column=self._variable_end_column, row=row).value = v
                    self._variable_end_column += 1
